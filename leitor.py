import os
from pdf2image import convert_from_path
import pytesseract
from PIL import Image
from openai_api import gerar_resposta
import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import sys
import re
from datetime import datetime

base_path = os.path.dirname(os.path.abspath(__file__))

# Tesseract local
pytesseract.pytesseract.tesseract_cmd = os.path.join(base_path, "Tesseract-OCR", "tesseract.exe")

resultados = []

caminho_arquivo_excel = os.path.join(base_path, "ProtocoloNotaFiscal.xlsx") # Caminho para a planilha Excel

arquivo = load_workbook(caminho_arquivo_excel) # Carrega a planilha

nome_ultima_aba = arquivo.sheetnames[-1] # Pega o nome da última aba como modelo de referencia
aba_origem = arquivo[nome_ultima_aba]

hoje = datetime.today()
dia = hoje.day
mes = hoje.month
ano = hoje.year

nova_aba_nome = f"{dia}-{mes}-{ano}"
arquivo.copy_worksheet(aba_origem).title = nova_aba_nome # Cria uma cópia do modelo de protocolo e altera o nome pra evitar conflito
aba_nova = arquivo[nova_aba_nome]

# Limpa as células antigas a partir da linha 3 pra garantir que o protocolo estará vazio
for row in aba_nova.iter_rows(min_row=3): 
    for cell in row:
        cell.value = None

def formatar_valor(valor):
    try:
        valor_float = float(valor.replace(",", "."))
        return f"R$ {valor_float:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    except:
        return valor  

pasta_notas = os.path.join(base_path, "NotasFiscais")

for nome_arquivo in os.listdir(pasta_notas):
    if nome_arquivo.lower().endswith(".pdf"):
        caminho_arquivo = os.path.join(pasta_notas, nome_arquivo)
        print(f"\n🔍 Lendo os textos do arquivo: {nome_arquivo}")
        
        try:
            # Poppler local
            poppler_path = os.path.join(base_path, "bin")  # Dentro dessa pasta estão os executáveis .exe
            paginas_imagem = convert_from_path(caminho_arquivo, poppler_path=poppler_path)

            texto_total = ""
            for pagina in paginas_imagem:
                texto_pagina = pytesseract.image_to_string(pagina) # Lê os textos dos arquivos e os armazena em uma variável
                texto_total += texto_pagina + "\n"

            print("📄 Texto analisado. Extraindo as informações...")

            prompt_unico = f"""
            O texto abaixo foi extraído de uma nota fiscal via OCR de uma nota fiscal. Ele pode conter erros de ortografia, gramática e formatação.

            Sua tarefa é:

            1. Ler e **corrigir mentalmente** o texto para entender seu conteúdo.
            2. Com base nisso, **extrair as seguintes informações**, e retorná-las em formato JSON:

            - "fornecedor": nome da empresa que emitiu a nota
            - "material": nome do produto que foi comprado ou caso houver varios produtos, retorne uma categoria na qual os produtos se enquadrem, caso não consiga retorne 'DIVERSOS'
            - "num_nf": numero da nota fiscal e da série, separados por '-' e tudo junto sem espaços.
            - "dt_emissao": quando a nota foi gerada, no formato DD/MM/AAAA
            - "data_vencimento": data de vencimento da nota, no formato DD/MM/AAAA
            - "valor_frete": retorne o valor do frete, somente se houver caso contrário, retorne 'SEM FRETE'
            - "valor_total_nota": valor total da nota, **'R$'**, com milhar separado por ponto (.) e decimais por vírgula (,)

            Se não encontrar alguma informação, preencha com '-'. ⚠️ Retorne **apenas** um JSON com exatamente essas 4 chaves e seus respectivos valores.

            **Texto a ser corrigido:**
            {texto_total}
            """

            resposta = gerar_resposta(prompt_unico)
            resposta_bruta = resposta["resposta"]
            resposta_limpa = re.sub(r"```json|```", "", resposta_bruta).strip()

            dados_ia = json.loads(resposta_limpa)

            resultados.append(dados_ia)

        except Exception as e:
            print(f"⚠️ Erro ao processar {nome_arquivo}: {e}")

df = pd.DataFrame(resultados, index=range(1, len(resultados) + 1))

df["valor_total_nota"] = df["valor_total_nota"].apply(formatar_valor)
ordem_colunas = ["fornecedor", "material", "num_nf", "dt_emissao", "data_vencimento", "valor_frete", "valor_total_nota"]
df = df[ordem_colunas]

# Escreve o DataFrame na aba a partir da linha 3
for i, row in enumerate(dataframe_to_rows(df, index=True, header=False), start=2):
    for j, value in enumerate(row, start=1):  # começa na coluna A
        aba_nova.cell(row=i, column=j, value=value)

arquivo.save(caminho_arquivo_excel)

input("\n\nPressione Enter para sair...")
sys.exit()
